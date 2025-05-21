import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from scipy.optimize import curve_fit

df = openpyxl.load_workbook("F60_motorSpec.xlsx")

df1 = df.active

motorThrottle_list = []
motorThrust_list = []
motorCurrent_list = []
motorRPM_list = []
motorPower_list = []


# for row in range(0, df1.max_row):
#     for col in df1.iter_cols(1, df1.max_column):
#         print(col[row].value)


''' Get thrust values '''
def getThrottle():
    for row in range(2, df1.max_row):
        for col in df1.iter_cols(1,1):
            # print(col[row].value)
            motorPower_list.append(col[row].value)
    return motorPower_list


def getThrust():
    for row in range(2, df1.max_row):
        for col in df1.iter_cols(2,2):
            # print(col[row].value)
            motorThrust_list.append(col[row].value)
    return motorThrust_list


def getCurrent():
    for row in range(2, df1.max_row):
        for col in df1.iter_cols(4,4):
            # print(col[row].value)
            motorCurrent_list.append(col[row].value)
    return motorCurrent_list


def getRPM():
    for row in range(2, df1.max_row):
        for col in df1.iter_cols(5,5):
            # print(col[row].value)
            motorRPM_list.append(col[row].value)
    return motorRPM_list


def getPower():
    for row in range(2, df1.max_row):
        for col in df1.iter_cols(6,6):
            # print(col[row].value)
            motorPower_list.append(col[row].value)
    return motorPower_list


def plotter():
    def linear(x, m, c):
        y = m*x + c
        return y
    
    # plot_ylist = getThrottle()
    # plot_ylist = getThrottle()
    plot_ylist = getThrust()
    plot_xlist = getCurrent()

    xdata = np.asarray(plot_xlist)
    ydata = np.asarray(plot_ylist)

    params, covariance = curve_fit(linear, xdata, ydata)

    fit_m = params[0]
    fit_c = params[1]

    fit_y = linear(xdata, fit_m, fit_c)

    plt.plot(xdata, ydata, label="data")
    plt.plot(xdata, fit_y, label="fit")
    plt.grid(color='k', linestyle='-', linewidth=0.5)
    plt.legend()
    plt.show()
    return fit_m, fit_c



def main():
    # getThrottle()
    # getThrust()
    # getCurrent()
    # getRPM()
    # getPower()
    plotter()


if __name__ == "__main__":
    main()
    